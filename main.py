import customtkinter as ctk
import numpy as np
import matplotlib.pyplot as plt
from scipy import signal
import os
import csv
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook, load_workbook
from PIL import Image

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


def aduna_polinomii(p1, p2):
    p1, p2 = np.array(p1), np.array(p2)

    if len(p1) < len(p2):
        p1 = np.pad(p1, (len(p2) - len(p1), 0))
    elif len(p2) < len(p1):
        p2 = np.pad(p2, (len(p1) - len(p2), 0))

    return (p1 + p2).tolist()


def aplica_pid(num_plant, den_plant, kp, ki, kd):
    num_pid = [kd, kp, ki]
    den_pid = [1, 0]

    num_open = np.polymul(num_pid, num_plant).tolist()
    den_open = np.polymul(den_pid, den_plant).tolist()

    num_closed = num_open
    den_closed = aduna_polinomii(den_open, num_open)

    return num_closed, den_closed


def calculeaza_indicatori_treapta(t, y):
    valoare_finala = y[-1]

    if abs(valoare_finala) > 1e-6:
        overshoot = ((np.max(y) - valoare_finala) / valoare_finala) * 100
    else:
        overshoot = 0

    prag = 0.02 * abs(valoare_finala)
    timp_stabilizare = t[-1]

    for i in range(len(y)):
        if np.all(np.abs(y[i:] - valoare_finala) <= prag):
            timp_stabilizare = t[i]
            break

    return valoare_finala, overshoot, timp_stabilizare


def calculeaza_indicatori_impuls(t, y):
    return np.max(y), np.min(y), t[np.argmax(y)]


def calculeaza_indicatori_rampa(t, y):
    valoare_finala = y[-1]
    valoare_maxima = np.max(y)
    eroare_finala = t[-1] - valoare_finala
    return valoare_finala, valoare_maxima, eroare_finala


def obtine_raspuns(sistem, t, tip_semnal):
    if tip_semnal == "Treapta":
        return signal.step(sistem, T=t)
    elif tip_semnal == "Impuls":
        return signal.impulse(sistem, T=t)
    else:
        u = t
        tout, y, _ = signal.lsim(sistem, U=u, T=t)
        return tout, y


def salveaza_grafic():
    os.makedirs("rezultate", exist_ok=True)
    nume_fisier = datetime.now().strftime("rezultate/grafic_%Y%m%d_%H%M%S.png")
    plt.savefig(nume_fisier, dpi=300, bbox_inches="tight")
    return nume_fisier


def salveaza_rezultate(rezultate):
    os.makedirs("rezultate", exist_ok=True)
    nume_fisier = datetime.now().strftime("rezultate/rezultate_%Y%m%d_%H%M%S.txt")

    with open(nume_fisier, "w", encoding="utf-8") as fisier:
        fisier.write(rezultate)

    return nume_fisier


def salveaza_istoric(date_istoric):
    os.makedirs("rezultate", exist_ok=True)
    nume_fisier = "rezultate/istoric_simulari.csv"
    fisier_exista = os.path.exists(nume_fisier)

    with open(nume_fisier, "a", newline="", encoding="utf-8") as fisier:
        writer = csv.writer(fisier)

        if not fisier_exista:
            writer.writerow([
                "Data", "Tip analiza", "Tip sistem", "Tip semnal",
                "Mod control", "Parametri", "Indicator 1",
                "Indicator 2", "Indicator 3"
            ])

        writer.writerow(date_istoric)

    return nume_fisier


def salveaza_istoric_excel(date_istoric):
    os.makedirs("rezultate", exist_ok=True)
    nume_fisier = "rezultate/istoric_simulari.xlsx"

    if os.path.exists(nume_fisier):
        workbook = load_workbook(nume_fisier)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Istoric simulari"
        sheet.append([
            "Data", "Tip analiza", "Tip sistem", "Tip semnal",
            "Mod control", "Parametri", "Indicator 1",
            "Indicator 2", "Indicator 3"
        ])

    sheet.append(date_istoric)
    workbook.save(nume_fisier)

    return nume_fisier


def salveaza_istoric_xml(date_istoric):
    os.makedirs("rezultate", exist_ok=True)
    nume_fisier = "rezultate/istoric_simulari.xml"

    campuri = [
        "Data",
        "TipAnaliza",
        "TipSistem",
        "TipSemnal",
        "ModControl",
        "Parametri",
        "Indicator1",
        "Indicator2",
        "Indicator3"
    ]

    if os.path.exists(nume_fisier):
        tree = ET.parse(nume_fisier)
        root = tree.getroot()
    else:
        root = ET.Element("IstoricSimulari")
        tree = ET.ElementTree(root)

    simulare = ET.SubElement(root, "Simulare")

    for camp, valoare in zip(campuri, date_istoric):
        element = ET.SubElement(simulare, camp)
        element.text = str(valoare)

    tree.write(
        nume_fisier,
        encoding="utf-8",
        xml_declaration=True
    )

    return nume_fisier


def construieste_rezultate(tip_sistem, tip_semnal, mod_control, parametri, t, y):
    if tip_semnal == "Treapta":
        vf, os, ts = calculeaza_indicatori_treapta(t, y)

        rezultate = (
            f"Tip simulare: {tip_sistem}\n"
            f"Tip semnal: {tip_semnal}\n"
            f"Mod control: {mod_control}\n"
            f"Parametri: {parametri}\n"
            f"Valoare finala: {vf:.2f}\n"
            f"Overshoot: {os:.2f}%\n"
            f"Timp stabilizare: {ts:.2f} s"
        )

        indicatori = [
            f"Valoare finala={vf:.2f}",
            f"Overshoot={os:.2f}%",
            f"Timp stabilizare={ts:.2f}s"
        ]

    elif tip_semnal == "Impuls":
        vmax, vmin, tmax = calculeaza_indicatori_impuls(t, y)

        rezultate = (
            f"Tip simulare: {tip_sistem}\n"
            f"Tip semnal: {tip_semnal}\n"
            f"Mod control: {mod_control}\n"
            f"Parametri: {parametri}\n"
            f"Valoare maxima: {vmax:.2f}\n"
            f"Valoare minima: {vmin:.2f}\n"
            f"Timp valoare maxima: {tmax:.2f} s"
        )

        indicatori = [
            f"Valoare maxima={vmax:.2f}",
            f"Valoare minima={vmin:.2f}",
            f"Timp maxim={tmax:.2f}s"
        ]

    else:
        vf, vmax, eroare = calculeaza_indicatori_rampa(t, y)

        rezultate = (
            f"Tip simulare: {tip_sistem}\n"
            f"Tip semnal: {tip_semnal}\n"
            f"Mod control: {mod_control}\n"
            f"Parametri: {parametri}\n"
            f"Valoare finala iesire: {vf:.2f}\n"
            f"Valoare maxima iesire: {vmax:.2f}\n"
            f"Eroare finala fata de rampa: {eroare:.2f}"
        )

        indicatori = [
            f"Valoare finala iesire={vf:.2f}",
            f"Valoare maxima iesire={vmax:.2f}",
            f"Eroare finala={eroare:.2f}"
        ]

    return rezultate, indicatori


def creeaza_sistem(num, den):
    if option_pid.get() == "Cu PID":
        kp = float(entry_kp.get())
        ki = float(entry_ki.get())
        kd = float(entry_kd.get())

        num, den = aplica_pid(num, den, kp, ki, kd)
        mod_control = f"Cu PID (Kp={kp}, Ki={ki}, Kd={kd})"
    else:
        mod_control = "Fara PID"

    return signal.TransferFunction(num, den), mod_control


def construieste_model_baza(tip_sistem):
    if tip_sistem == "Sistem ordin I":
        K = float(entry_k.get())
        T = float(entry_t.get())
        return [K], [T, 1], f"K={K}, T={T}"

    elif tip_sistem == "Sistem ordin II":
        wn = float(entry_wn.get())
        zeta = float(entry_zeta.get())
        return [wn ** 2], [1, 2 * zeta * wn, wn ** 2], f"wn={wn}, zeta={zeta}"

    else:
        wn = float(entry_cmp_wn.get())
        zeta1 = float(entry_cmp_zeta1.get())
        return [wn ** 2], [1, 2 * zeta1 * wn, wn ** 2], f"wn={wn}, zeta={zeta1}"


def afiseaza_status(text):
    status_box.configure(state="normal")
    status_box.delete("1.0", "end")
    status_box.insert("1.0", text)
    status_box.configure(state="disabled")


def incarca_istoric():
    fisier = "rezultate/istoric_simulari.csv"

    istoric_box.configure(state="normal")
    istoric_box.delete("1.0", "end")

    if not os.path.exists(fisier):
        istoric_box.insert("1.0", "Nu există încă simulări salvate.")
        istoric_box.configure(state="disabled")
        return

    with open(fisier, "r", encoding="utf-8") as f:
        reader = list(csv.reader(f))

    if len(reader) <= 1:
        istoric_box.insert("1.0", "Fișierul de istoric există, dar nu conține simulări.")
        istoric_box.configure(state="disabled")
        return

    header = reader[0]
    ultimele = reader[-10:]

    text = "Ultimele simulări salvate:\n\n"

    for index, rand in enumerate(ultimele, start=1):
        text += f"Simulare {index}\n"
        for h, valoare in zip(header, rand):
            text += f"{h}: {valoare}\n"
        text += "-" * 60 + "\n"

    istoric_box.insert("1.0", text)
    istoric_box.configure(state="disabled")


def sterge_istoric():
    fisiere = [
        "rezultate/istoric_simulari.csv",
        "rezultate/istoric_simulari.xlsx",
        "rezultate/istoric_simulari.xml"
    ]

    for fisier in fisiere:
        if os.path.exists(fisier):
            os.remove(fisier)

    incarca_istoric()
    afiseaza_status("Istoricul simulărilor a fost șters.")


def schimba_tema(choice):
    if choice == "Dark":
        ctk.set_appearance_mode("dark")
    else:
        ctk.set_appearance_mode("light")


def deschide_folder_rezultate():
    os.makedirs("rezultate", exist_ok=True)
    cale = os.path.abspath("rezultate")
    os.startfile(cale)


def finalizeaza_export(rezultate, date_istoric, text_pe_grafic=True):
    if text_pe_grafic:
        plt.xlabel("Timp [s]")
        plt.ylabel("Iesire")
        plt.grid(True)
        plt.legend()

        plt.figtext(
            0.12,
            0.72,
            rezultate,
            fontsize=8,
            bbox={"facecolor": "white", "alpha": 0.8}
        )
    else:
        plt.tight_layout()

    nume_grafic = salveaza_grafic()
    nume_rezultate = salveaza_rezultate(rezultate)
    nume_istoric = salveaza_istoric(date_istoric)
    nume_excel = salveaza_istoric_excel(date_istoric)
    nume_xml = salveaza_istoric_xml(date_istoric)

    mesaj = (
        rezultate
        + f"\n\nGrafic salvat:\n{nume_grafic}"
        + f"\n\nRezultate salvate:\n{nume_rezultate}"
        + f"\n\nCSV actualizat:\n{nume_istoric}"
        + f"\n\nExcel actualizat:\n{nume_excel}"
        + f"\n\nXML actualizat:\n{nume_xml}"
    )

    afiseaza_status(mesaj)
    incarca_istoric()

    plt.show()


def simuleaza_raspuns_timp():
    tip_sistem = option_sistem.get()
    tip_semnal = option_semnal.get()
    data_curenta = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    t = np.linspace(0, 10, 1000)
    plt.figure(figsize=(8, 5))

    if tip_sistem == "Comparare ordin II":
        wn = float(entry_cmp_wn.get())
        zeta1 = float(entry_cmp_zeta1.get())
        zeta2 = float(entry_cmp_zeta2.get())

        num1 = [wn ** 2]
        den1 = [1, 2 * zeta1 * wn, wn ** 2]

        num2 = [wn ** 2]
        den2 = [1, 2 * zeta2 * wn, wn ** 2]

        sistem1, mod_control = creeaza_sistem(num1, den1)
        sistem2, _ = creeaza_sistem(num2, den2)

        t, y1 = obtine_raspuns(sistem1, t, tip_semnal)
        t, y2 = obtine_raspuns(sistem2, t, tip_semnal)

        rezultate1, indicatori1 = construieste_rezultate(
            "Sistem 1", tip_semnal, mod_control, f"wn={wn}, zeta={zeta1}", t, y1
        )

        rezultate2, indicatori2 = construieste_rezultate(
            "Sistem 2", tip_semnal, mod_control, f"wn={wn}, zeta={zeta2}", t, y2
        )

        rezultate = (
            f"Tip analiza: Raspuns in timp\n"
            f"Tip simulare: Comparare sisteme ordin II\n"
            f"Tip semnal: {tip_semnal}\n"
            f"Mod control: {mod_control}\n"
            f"Parametri comuni: wn={wn}\n\n"
            f"{rezultate1}\n\n"
            f"{rezultate2}"
        )

        date_istoric = [
            data_curenta, "Raspuns in timp", "Comparare ordin II",
            tip_semnal, mod_control, f"wn={wn}, zeta1={zeta1}, zeta2={zeta2}",
            f"S1: {indicatori1[0]} | S2: {indicatori2[0]}",
            f"S1: {indicatori1[1]} | S2: {indicatori2[1]}",
            f"S1: {indicatori1[2]} | S2: {indicatori2[2]}"
        ]

        plt.plot(t, y1, label=f"Sistem 1: zeta={zeta1}")
        plt.plot(t, y2, label=f"Sistem 2: zeta={zeta2}")

        if tip_semnal == "Rampa":
            plt.plot(t, t, "--", label="Intrare rampa")

        plt.title(f"Comparatie raspuns la {tip_semnal.lower()} - Ordin II")

    else:
        num, den, parametri = construieste_model_baza(tip_sistem)
        sistem, mod_control = creeaza_sistem(num, den)

        t, y = obtine_raspuns(sistem, t, tip_semnal)

        rezultate, indicatori = construieste_rezultate(
            tip_sistem, tip_semnal, mod_control, parametri, t, y
        )

        rezultate = "Tip analiza: Raspuns in timp\n" + rezultate

        date_istoric = [
            data_curenta, "Raspuns in timp", tip_sistem, tip_semnal,
            mod_control, parametri, indicatori[0], indicatori[1], indicatori[2]
        ]

        plt.plot(t, y, label=f"{parametri} - {mod_control}")

        if tip_semnal == "Rampa":
            plt.plot(t, t, "--", label="Intrare rampa")

        plt.title(f"Raspuns la {tip_semnal.lower()} - {tip_sistem}")

    finalizeaza_export(rezultate, date_istoric)


def simuleaza_bode():
    tip_sistem = option_sistem.get()
    data_curenta = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    num, den, parametri = construieste_model_baza(tip_sistem)
    sistem, mod_control = creeaza_sistem(num, den)

    w, mag, phase = signal.bode(sistem)

    plt.figure(figsize=(9, 6))

    plt.subplot(2, 1, 1)
    plt.semilogx(w, mag)
    plt.title(f"Diagrama Bode - {tip_sistem}")
    plt.ylabel("Magnitudine [dB]")
    plt.grid(True, which="both")

    plt.subplot(2, 1, 2)
    plt.semilogx(w, phase)
    plt.xlabel("Frecventa [rad/s]")
    plt.ylabel("Faza [grade]")
    plt.grid(True, which="both")

    mag_max = np.max(mag)
    mag_min = np.min(mag)
    faza_min = np.min(phase)

    rezultate = (
        f"Tip analiza: Diagrama Bode\n"
        f"Tip sistem: {tip_sistem}\n"
        f"Mod control: {mod_control}\n"
        f"Parametri: {parametri}\n"
        f"Magnitudine maxima: {mag_max:.2f} dB\n"
        f"Magnitudine minima: {mag_min:.2f} dB\n"
        f"Faza minima: {faza_min:.2f} grade"
    )

    date_istoric = [
        data_curenta, "Diagrama Bode", tip_sistem, "-",
        mod_control, parametri,
        f"Magnitudine maxima={mag_max:.2f} dB",
        f"Magnitudine minima={mag_min:.2f} dB",
        f"Faza minima={faza_min:.2f} grade"
    ]

    finalizeaza_export(rezultate, date_istoric, text_pe_grafic=False)


def simuleaza():
    try:
        if option_analiza.get() == "Raspuns in timp":
            simuleaza_raspuns_timp()
        else:
            simuleaza_bode()

    except ValueError:
        afiseaza_status("Eroare: introdu valori numerice valide!")


def actualizeaza_campuri(choice):
    afiseaza_status("")
    frame_ordin_i.pack_forget()
    frame_ordin_ii.pack_forget()
    frame_comparare.pack_forget()

    if choice == "Sistem ordin I":
        frame_ordin_i.pack(pady=10)
    elif choice == "Sistem ordin II":
        frame_ordin_ii.pack(pady=10)
    else:
        frame_comparare.pack(pady=10)


def actualizeaza_pid(choice):
    afiseaza_status("")

    if choice == "Cu PID":
        frame_pid.pack(pady=10)
    else:
        frame_pid.pack_forget()


app = ctk.CTk()
app.title("Platforma Software pentru Simulare")
app.geometry("900x900")

if os.path.exists("logo.png"):
    logo_image = ctk.CTkImage(
        light_image=Image.open("logo.png"),
        dark_image=Image.open("logo.png"),
        size=(110, 110)
    )

    logo_label = ctk.CTkLabel(app, image=logo_image, text="")
    logo_label.pack(pady=(10, 0))

title = ctk.CTkLabel(
    app,
    text="Platforma pentru simularea sistemelor de control",
    font=("Arial", 24)
)
title.pack(pady=10)

tabview = ctk.CTkTabview(app, width=850, height=700)
tabview.pack(padx=20, pady=10)

tab_simulare = tabview.add("Simulare")
tab_istoric = tabview.add("Istoric")
tab_despre = tabview.add("Despre aplicație")


left_frame = ctk.CTkFrame(tab_simulare)
left_frame.pack(side="left", fill="y", padx=15, pady=15)

right_frame = ctk.CTkFrame(tab_simulare)
right_frame.pack(side="right", fill="both", expand=True, padx=15, pady=15)

ctk.CTkLabel(left_frame, text="Configurare simulare", font=("Arial", 18)).pack(pady=10)

ctk.CTkLabel(left_frame, text="Temă interfață").pack(pady=3)
option_tema = ctk.CTkOptionMenu(left_frame, values=["Dark", "Light"], command=schimba_tema)
option_tema.pack(pady=5)
option_tema.set("Dark")

ctk.CTkLabel(left_frame, text="Tip analiză").pack(pady=3)
option_analiza = ctk.CTkOptionMenu(left_frame, values=["Raspuns in timp", "Diagrama Bode"])
option_analiza.pack(pady=5)
option_analiza.set("Raspuns in timp")

ctk.CTkLabel(left_frame, text="Tip sistem").pack(pady=3)
option_sistem = ctk.CTkOptionMenu(
    left_frame,
    values=["Sistem ordin I", "Sistem ordin II", "Comparare ordin II"],
    command=actualizeaza_campuri
)
option_sistem.pack(pady=5)
option_sistem.set("Sistem ordin I")

ctk.CTkLabel(left_frame, text="Tip semnal de intrare").pack(pady=3)
option_semnal = ctk.CTkOptionMenu(left_frame, values=["Treapta", "Impuls", "Rampa"])
option_semnal.pack(pady=5)
option_semnal.set("Treapta")

ctk.CTkLabel(left_frame, text="Mod control").pack(pady=3)
option_pid = ctk.CTkOptionMenu(
    left_frame,
    values=["Fara PID", "Cu PID"],
    command=actualizeaza_pid
)
option_pid.pack(pady=5)
option_pid.set("Fara PID")


frame_ordin_i = ctk.CTkFrame(left_frame)
frame_ordin_i.pack(pady=10)

ctk.CTkLabel(frame_ordin_i, text="Amplificare K").pack(pady=5)
entry_k = ctk.CTkEntry(frame_ordin_i)
entry_k.pack(pady=5)
entry_k.insert(0, "2")

ctk.CTkLabel(frame_ordin_i, text="Constanta de timp T").pack(pady=5)
entry_t = ctk.CTkEntry(frame_ordin_i)
entry_t.pack(pady=5)
entry_t.insert(0, "1")


frame_ordin_ii = ctk.CTkFrame(left_frame)

ctk.CTkLabel(frame_ordin_ii, text="Frecvența naturală wn").pack(pady=5)
entry_wn = ctk.CTkEntry(frame_ordin_ii)
entry_wn.pack(pady=5)
entry_wn.insert(0, "3")

ctk.CTkLabel(frame_ordin_ii, text="Factor amortizare zeta").pack(pady=5)
entry_zeta = ctk.CTkEntry(frame_ordin_ii)
entry_zeta.pack(pady=5)
entry_zeta.insert(0, "0.3")


frame_comparare = ctk.CTkFrame(left_frame)

ctk.CTkLabel(frame_comparare, text="Frecvența naturală wn").pack(pady=5)
entry_cmp_wn = ctk.CTkEntry(frame_comparare)
entry_cmp_wn.pack(pady=5)
entry_cmp_wn.insert(0, "3")

ctk.CTkLabel(frame_comparare, text="Zeta sistem 1").pack(pady=5)
entry_cmp_zeta1 = ctk.CTkEntry(frame_comparare)
entry_cmp_zeta1.pack(pady=5)
entry_cmp_zeta1.insert(0, "0.2")

ctk.CTkLabel(frame_comparare, text="Zeta sistem 2").pack(pady=5)
entry_cmp_zeta2 = ctk.CTkEntry(frame_comparare)
entry_cmp_zeta2.pack(pady=5)
entry_cmp_zeta2.insert(0, "0.7")


frame_pid = ctk.CTkFrame(left_frame)

ctk.CTkLabel(frame_pid, text="Parametri PID").pack(pady=5)

ctk.CTkLabel(frame_pid, text="Kp").pack(pady=3)
entry_kp = ctk.CTkEntry(frame_pid)
entry_kp.pack(pady=3)
entry_kp.insert(0, "2")

ctk.CTkLabel(frame_pid, text="Ki").pack(pady=3)
entry_ki = ctk.CTkEntry(frame_pid)
entry_ki.pack(pady=3)
entry_ki.insert(0, "1")

ctk.CTkLabel(frame_pid, text="Kd").pack(pady=3)
entry_kd = ctk.CTkEntry(frame_pid)
entry_kd.pack(pady=3)
entry_kd.insert(0, "0.2")


button = ctk.CTkButton(left_frame, text="Simulează", command=simuleaza)
button.pack(pady=20)

ctk.CTkLabel(right_frame, text="Rezultatele ultimei simulări", font=("Arial", 18)).pack(pady=10)

status_box = ctk.CTkTextbox(right_frame, width=450, height=500)
status_box.pack(padx=10, pady=10, fill="both", expand=True)
status_box.configure(state="disabled")


ctk.CTkLabel(tab_istoric, text="Istoric simulări", font=("Arial", 22)).pack(pady=15)

ctk.CTkLabel(
    tab_istoric,
    text="Aici sunt afișate ultimele simulări salvate în fișierul CSV.",
    font=("Arial", 14)
).pack(pady=5)

ctk.CTkButton(tab_istoric, text="Actualizează istoric", command=incarca_istoric).pack(pady=5)

ctk.CTkButton(
    tab_istoric,
    text="Deschide folder rezultate",
    command=deschide_folder_rezultate
).pack(pady=5)

ctk.CTkButton(
    tab_istoric,
    text="Șterge istoric",
    fg_color="darkred",
    hover_color="red",
    command=sterge_istoric
).pack(pady=5)

istoric_box = ctk.CTkTextbox(tab_istoric, width=780, height=450)
istoric_box.pack(padx=20, pady=10)
istoric_box.configure(state="disabled")


ctk.CTkLabel(
    tab_despre,
    text="Platformă software pentru simularea și analiza unui sistem de control automat",
    font=("Arial", 20),
    wraplength=750
).pack(pady=25)

despre_text = (
    "Această aplicație a fost realizată pentru simularea și analiza sistemelor de control automat.\n\n"
    "Funcționalități principale:\n"
    "- simularea sistemelor de ordin I și ordin II;\n"
    "- analiza răspunsului la treaptă, impuls și rampă;\n"
    "- compararea a două sisteme de ordin II;\n"
    "- utilizarea unui controler PID simplu;\n"
    "- generarea diagramei Bode;\n"
    "- calcularea automată a indicatorilor specifici;\n"
    "- exportul rezultatelor în PNG, TXT, CSV, Excel și XML;\n"
    "- afișarea istoricului simulărilor direct în interfață;\n"
    "- deschiderea automată a folderului de rezultate;\n"
    "- comutare între tema Dark și Light;\n"
    "- ștergerea istoricului de simulări;\n"
    "- integrarea unui logo grafic în interfață.\n\n"
    "Tehnologii utilizate:\n"
    "- Python;\n"
    "- CustomTkinter pentru interfața grafică;\n"
    "- NumPy pentru calcule numerice;\n"
    "- SciPy pentru modelarea și simularea sistemelor;\n"
    "- Matplotlib pentru reprezentări grafice;\n"
    "- OpenPyXL pentru exportul în Excel;\n"
    "- XML ElementTree pentru exportul în XML;\n"
    "- Pillow pentru încărcarea logo-ului."
)

ctk.CTkLabel(
    tab_despre,
    text=despre_text,
    font=("Arial", 15),
    justify="left",
    wraplength=760
).pack(padx=30, pady=20)

incarca_istoric()

app.mainloop()